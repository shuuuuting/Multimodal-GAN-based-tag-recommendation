#%%----------------------------------------------------------------------------------------
#取出 text&tag資料 / split id
#------------------------------------------------------------------------------------------
import os
import numpy as np
import pandas as pd
import csv
from pymongo import MongoClient

conn = MongoClient('localhost', 27017) #連結mongodb
db = conn.NiusNews2020_04_12 #create database
text_train = pd.DataFrame(list(db['text_train_db'].find({},{"_id": 0,"postid": 1,"content_jieba": 1})))
text_test = pd.DataFrame(list(db['text_test_db'].find({},{"_id": 0,"postid": 1,"content_jieba": 1})))
tag_train = np.array(pd.DataFrame(list(db['tag_train_db'].find({},{"_id": 0}))))
tag_test = np.array(pd.DataFrame(list(db['tag_test_db'].find({},{"_id": 0}))))
train_id = db['split_indices'].find_one({"dataname":"train"})['indexlist']
test_id = db['split_indices'].find_one({"dataname":"test"})['indexlist']


#%%----------------------------------------------------------------------------------------
#preprocess text for embedding layer
#------------------------------------------------------------------------------------------
from keras.preprocessing.sequence import pad_sequences
from keras.preprocessing.text import one_hot
text_train['sentence'] = [' '.join(map(str, l)) for l in text_train['content_jieba']]  
text_test['sentence'] = [' '.join(map(str, l)) for l in text_test['content_jieba']]  

#backup data
text_tr_pd = text_train
text_te_pd = text_test

doc_train = list(text_train['sentence'])
doc_test = list(text_test['sentence'])

#count no of vocs in corpus
from collections import Counter 
voc_list = [item for sublist in text_train['content_jieba'] for item in sublist]   
voc_list.extend([item for sublist in text_test['content_jieba'] for item in sublist])   
vocs = Counter(voc_list).keys() 
print("No of unique vocs are:", len(vocs)) #45807

#integer encode the documents
vocab_size = len(vocs)
text_train = [one_hot(d, vocab_size) for d in doc_train]
text_test = [one_hot(d, vocab_size) for d in doc_test]

#pad documents to a max length of 50 words
max_length = 50
text_train = pad_sequences(text_train, maxlen=max_length, padding='post')
print("padded_docs_train:\n",text_train)
text_test = pad_sequences(text_test, maxlen=max_length, padding='post')

#%%----------------------------------------------------------------------------------------
#extract image features using VGG16 
#------------------------------------------------------------------------------------------
from keras.preprocessing.image import load_img
from keras.preprocessing.image import img_to_array
from keras.applications.vgg16 import preprocess_input
from keras.applications.vgg16 import decode_predictions
from keras.applications.vgg16 import VGG16
from keras.models import Model
import cv2
from PIL import Image
from gridfs import *
import io
gridFS = GridFS(db, collection="fs")
#load the model
base_model = VGG16(weights='imagenet')
model = Model(inputs=base_model.input, outputs=base_model.get_layer('block5_pool').output) #拿這層的output當成圖片feature
image_train = []
count = 0
print('Processing train image:')
#load image from db
for idx in train_id:
    if count%20==0:
        print(count)
    count+=1
    image_data = gridFS.find_one({"filename": str(idx)})
    #convert the image pixels to a numpy array
    image = np.array(Image.open(io.BytesIO(image_data.read())).convert('RGB'))
    image = cv2.resize(image, dsize=(224, 224), interpolation=cv2.INTER_CUBIC)
    #img = Image.fromarray(image)
    #img.show()
    #reshape data for the model
    image = image.reshape((1, image.shape[0], image.shape[1], image.shape[2]))
    #image = image/255.0
    # prepare the image for the VGG model
    image = preprocess_input(image)
    # predict the probability across all output classes
    feature = model.predict(image)
    image_train.append(feature)
image_train = np.vstack(image_train)

image_test = []
count = 0
print('Processing test image:')
#load image from db
for idx in test_id:
    if count%20==0:
        print(count)
    count+=1
    image_data = gridFS.find_one({"filename": str(idx)})
    #convert the image pixels to a numpy array
    image = np.array(Image.open(io.BytesIO(image_data.read())).convert('RGB'))
    image = cv2.resize(image, dsize=(224, 224), interpolation=cv2.INTER_CUBIC)
    #img = Image.fromarray(image)
    #img.show()
    #reshape data for the model
    image = image.reshape((1, image.shape[0], image.shape[1], image.shape[2]))
    #image = image/255.0
    # prepare the image for the VGG model
    image = preprocess_input(image)
    # predict the probability across all output classes
    feature = model.predict(image)
    image_test.append(feature)
image_test = np.vstack(image_test)

'''
# convert the probabilities to class labels
label = decode_predictions(feature)
# retrieve the most likely result, e.g. highest probability
label = label[0][0]
# print the classification
print('%s (%.2f%%)' % (label[1], label[2]*100))
'''

#%%----------------------------------------------------------------------------------------
#seldDef.py
#------------------------------------------------------------------------------------------
import keras.backend as K
from keras.engine.topology import Layer, InputSpec
class coAttention_alt(Layer):
    """
    alternative co-attention
    inputs: [image feature tensor, hidden text feature tensor]
    output: co-Attention feature of image and text
    input dimensions:[(batchSize, num_region, CNN_dimension),
                    (batchSize, seq_length, CNN_dimension)]
    output dimension: batch_size*CNN_dimension
    """
    def __init__(self, dim_k, name='co_attn_layer', **kwargs):
        self.dim_k = dim_k  # internal tensor dimension
        # self.input_spec = InputSpec(min_ndim=3)
        self.supports_masking = True
        super(coAttention_alt, self).__init__(name=name)

    def build(self, input_shape): #build:定義權重的function
        if not isinstance(input_shape, list): #input_shape要是list
            raise ValueError('A Co-Attention_alt layer should be called '
                             'on a list of inputs.')
        if len(input_shape) != 2: #input應該包含兩個元件:img,text
            raise ValueError('A Co-Attention_alt layer should be called on a list of 2 inputs.'
                             'Got '+str(len(input_shape))+'inputs.')
        # print(input_shape)
        self.num_imgRegion = input_shape[0][1]
        self.seq_len = input_shape[1][1]
        self.output_dim = input_shape[0][2]

        self.w_Dense_Vi_0 = self.add_weight(name='w_Dense_Vi_0',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Vt_0 = self.add_weight(name='w_Dense_Vt_0',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Pi_0 = self.add_weight(name='w_Dense_Pi_0',
                                            shape=(2*self.dim_k, 1),
                                            initializer='random_normal',
                                            trainable=True)
        self.b_Dense_Pi_0 = self.add_weight(name='b_Dense_Pi_0',
                                            shape=(self.num_imgRegion,),
                                            initializer='zeros',
                                            trainable=True)

        self.w_Dense_Vi_1 = self.add_weight(name='w_Dense_Vi_1',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Vt_1 = self.add_weight(name='w_Dense_Vt_1',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Pi_1 = self.add_weight(name='w_Dense_Pi_1',
                                            shape=(2*self.dim_k, 1),
                                            initializer='random_normal',
                                            trainable=True)
        self.b_Dense_Pi_1 = self.add_weight(name='b_Dense_Pi_1',
                                            shape=(self.seq_len,),
                                            initializer='zeros',
                                            trainable=True)

        super(coAttention_alt, self).build(input_shape)  # Be sure to call this somewhere! 一定要在最後調用它!

    def call(self, x, mask=None): #call:編寫layer邏輯的function，執行forward propagation
        ifeature = x[0]
        tfeature_h = x[1]
        # tfeature = x[2]
        output_dim = self.output_dim
        num_imgRegion = self.num_imgRegion
        dim_k = self.dim_k
        seq_len = self.seq_len
        tfeature = K.mean(tfeature_h, axis=1)

        # phase 0: text-guided image feature computation
        w_Vi_0 = K.dot(K.reshape(ifeature, [-1, output_dim]), self.w_Dense_Vi_0) # shape=((batchSize*num_imgRegion),dim_k)
        w_Vi_0 = K.reshape(w_Vi_0, [-1, num_imgRegion, dim_k])  # shape=(batchSize,num_imgRegion,dim_k)
        w_Vt_0 = K.repeat(K.dot(tfeature, self.w_Dense_Vt_0), num_imgRegion)  # shape=(batchSize,num_imgRegion,dim_k) #未repeat前的dim是(batchSize,dim_k)
        Vi_Vt_0 = K.concatenate([w_Vi_0, w_Vt_0], axis=-1)  # shape=(batchSize,num_imgRegion,2*dim_k)
        Hi = K.tanh(Vi_Vt_0)
        # Hi_w = K.squeeze(K.dot(K.reshape(Hi, [-1, 2*dim_k]), self.w_Dense_Pi_0), axis=-1)
        # Hi_w_b = K.reshape(Hi_w, [-1, num_imgRegion]) + self.b_Dense_Pi_0
        Hi_w_b = K.squeeze(K.dot(Hi, self.w_Dense_Pi_0), axis=-1) + self.b_Dense_Pi_0  # shape=(batchSize,num_imgRegion) #axis是要丟棄的軸 #squeeze是用來刪掉維數是1的維度
        Pi = K.softmax(Hi_w_b) # shape=(batchSize,num_imgRegion)
        Pi = K.permute_dimensions(K.repeat(Pi, output_dim), (0, 2, 1))  # shape=(batchSize,num_imgRegion,output_dim)
        Pi_Vi = Pi*ifeature # shape=(batchSize,num_imgRegion,output_dim)*(batchSize,num_imgRegion,output_dim)
        Vi = K.sum(Pi_Vi, axis=1)  # shape=(batchSize,output_dim)

        # phase 1: image-guided text feature computation
        w_Vi_1 = K.repeat(K.dot(Vi, self.w_Dense_Vi_1), seq_len)    # shape=(batchSize,seq_len,dim_k)
        w_Vt_1 = K.dot(K.reshape(tfeature_h, [-1, output_dim]), self.w_Dense_Vt_1)   # shape=((batchSize*seq_len),dim_k)
        w_Vt_1 = K.reshape(w_Vt_1, (-1, seq_len, dim_k))    # shape= (batchSize, seq_len, dim_k)
        Vi_Vt_1 = K.concatenate([w_Vi_1, w_Vt_1], axis=-1)    # shape=(batchSize, seq_len, 2*dim_k)
        Ht = K.tanh(Vi_Vt_1)
        Ht_b = K.squeeze(K.dot(Ht, self.w_Dense_Pi_1), axis=-1) + self.b_Dense_Pi_1   # shape=(batch_size, seq_len)
        Pt = K.softmax(Ht_b)
        Pt = K.permute_dimensions(K.repeat(Pt, output_dim), (0, 2, 1))    # shape=(batchSize, seq_len, output_dim)
        Pt_Vt = Pt*tfeature_h # shape=(batchSize,seq_len,output_dim)*(batchSize,seq_len,output_dim)
        Vt = K.sum(Pt_Vt, axis=1)    # shape=(batchSize, output_dim)
        return Vi+Vt
        #return K.concatenate([Vi,Vt], axis=-1)

    def compute_output_shape(self, input_shape):
        output_shape = (input_shape[0][0], input_shape[0][-1]) #(batch_size,CNN_dimension)
        return output_shape

    def get_config(self):
        return super(coAttention_alt, self).get_config()

def myLossFunc(y_true, y_pred):
    probs_log = -K.log(y_pred) #此算法可能造成y_pred越接近1越好
    loss = K.mean(K.sum(probs_log*y_true, axis=-1))
    # loss = K.mean(K.sum(K.clip(probs_log * y_true, -1e40, 100), axis=-1))
    return loss
   
def myBinaryLossFunc(y_true, y_pred):
    probs_log = -K.log(y_pred)
    minus_log = -K.log(1-y_pred)
    loss = K.mean(K.sum(probs_log*y_true + minus_log*(1-y_true), axis=-1))
    return loss

#%%(沒用到)
labels = pd.DataFrame(list(db['labels'].find({})))
labels = labels['labels']
labels = [i[:-1] for i in labels]
labels_w2v = []
for label in labels:
    try:
        labels_w2v.append(w2vmodel.wv[label])
    except:
        labels_w2v.append(np.zeros(200))
labels_w2v = np.array(labels_w2v)

labelsf = ' '.join(map(str, labels))
labelsf = one_hot(labelsf, vocab_size)

#%%----------------------------------------------------------------------------------------
#co-attention.py 
#------------------------------------------------------------------------------------------
from keras.models import Model
from keras.layers.core import Activation, Flatten, Reshape, RepeatVector
from keras.layers.recurrent import LSTM
from keras.layers import Bidirectional
from keras.layers.convolutional import AveragePooling1D,Conv1D
from keras.layers.wrappers import TimeDistributed
from keras.layers import Input, Dense, Embedding, Dropout, Lambda, Add

num_tags = tag_train.shape[1]
num_words = vocab_size
#index_from = 3
seq_length = max_length
batch_size = 64
embedding_size = 200
hidden_size = 100
attention_size = 200
dim_k = 100
num_region = 7*7
drop_rate = 0.2
TopK = 5
shared_embed = Embedding(input_dim=num_words, output_dim=embedding_size,
                           mask_zero=True, input_length=seq_length)

def imageFeature(inputs):
    features = Reshape(target_shape=(num_region, 512))(inputs)
    features = Dense(embedding_size, activation="tanh", use_bias=False)(features) #single layer to convert each img vector into a new same dim vector as text feature vector
    features = SeqSelfAttention(attention_activation='sigmoid')(features)
    features_pooling = AveragePooling1D(pool_size=num_region, padding="same")(features)
    features_pooling = Lambda(lambda x: K.squeeze(x, axis=1))(features_pooling)

    return features, features_pooling

def textFeature(X):
    #X = Embedding(input_dim=num_words, output_dim=embedding_size,
    #                       mask_zero=True, input_length=seq_length)(X) #seq_length:一次輸入带有的詞彙個數
    #X = shared_embed(X)
    #tFeature = LSTM(units=embedding_size, return_sequences=True)(embeddings)
    #tFeature = Bidirectional(LSTM(units=embedding_size, return_sequences=True), merge_mode='sum')(embeddings)
    tFeature1 = Conv1D(embedding_size, 3, padding='same', strides=1, activation='relu')(X)
    tFeature2 = Conv1D(embedding_size, 4, padding='same', strides=1, activation='relu')(X)
    tFeature3 = Conv1D(embedding_size, 5, padding='same', strides=1, activation='relu')(X)
    tFeature = Add()([tFeature1, tFeature2, tFeature3])
    tFeature = SeqSelfAttention(attention_activation='sigmoid')(tFeature)

    return tFeature

def labelFeature(X):
    #X = Embedding(input_dim=num_words, output_dim=embedding_size,
    #                       mask_zero=True, input_length=seq_length)(X) #seq_length:一次輸入带有的詞彙個數
    #X = shared_embed(X)
    #tFeature = LSTM(units=embedding_size, return_sequences=True)(embeddings)
    #tFeature = Bidirectional(LSTM(units=embedding_size, return_sequences=True), merge_mode='sum')(embeddings)
    tFeature1 = Conv1D(embedding_size, 3, padding='same', strides=1, activation='relu')(X)
    tFeature2 = Conv1D(embedding_size, 4, padding='same', strides=1, activation='relu')(X)
    tFeature3 = Conv1D(embedding_size, 5, padding='same', strides=1, activation='relu')(X)
    tFeature = Add()([tFeature1, tFeature2, tFeature3])
    tFeature = SeqSelfAttention(attention_activation='sigmoid')(tFeature)

    return tFeature

def modelDef():
    inputs_img = Input(shape=(7, 7, 512))
    inputs_text = Input(shape=(seq_length,embedding_size)) #+title_length
    iFeature, iFeature_pooling = imageFeature(inputs_img)
    tFeature = textFeature(inputs_text)
    iFeature.set_shape((inputs_img.shape[0],num_region,embedding_size))
    tFeature.set_shape((inputs_text.shape[0],seq_length,embedding_size)) #+title_length
    #labelfeature =  shared_embed(np.array(labelsf))
    #labelfeature = labelFeature(tf.expand_dims(np.array(labelsf), axis=0))

    co_feature = coAttention_alt(dim_k=dim_k)([iFeature, tFeature])
    #co_feature = Attention()([co_feature, K.squeeze(labelfeature,axis=0)])
    dropout = Dropout(drop_rate)(co_feature)
    Softmax = Dense(num_tags, activation="softmax", use_bias=True)(dropout)
    Sigmoid = Dense(num_tags, activation="sigmoid", use_bias=True)(dropout)
    model = Model(inputs=[inputs_img, inputs_text],
                  outputs=[Sigmoid])
    optimizer = Adam(lr=0.0005, beta_1=0.5, decay=8e-8)
    # adam = optimizers.adam(lr=0.001, beta_1=0.9, beta_2=0.999, epsilon=1e-08, decay=0.00001)
    model.compile(optimizer='adam', loss='BinaryCrossentropy')
    #model.compile(optimizer='adam', loss=myLossFunc)
    return model

#%%----------------------------------------------------------------------------------------
#run co-attention.py 
#------------------------------------------------------------------------------------------
from openpyxl import load_workbook
from keras.callbacks import EarlyStopping
method = 'CoA'
wb = load_workbook(resultdir+'evaluation2.xlsx')
ws = wb.create_sheet(title=method)
ws.append(['acc', 'precision', 'recall', 'f1', 'ndcg', 'map'])
if not os.path.exists(resultdir+method):
    os.makedirs(resultdir+method)

early_stopping = EarlyStopping(monitor='val_loss', patience=10, verbose=1, mode='min', restore_best_weights=True) #持續5個epoch沒下降就停
round = 10
for i in range(round):
    print('round', i+1)
    model = modelDef()
    history = model.fit(x=[image_train, embedding_train], #embedding_train text_title_train
                        y=tag_train.astype(np.float32),
                        batch_size=batch_size,
                        epochs=50,
                        verbose=1,)
                    #validation_data=([image_test, text_test], tag_test.astype(np.float32)), 
                    #callbacks=[early_stopping])
    #model.save_weights('pretrained_G_w.h5')
    y_pred = model.predict(x=[image_test, embedding_test]) #embedding_test text_title_test
    acc_K, precision_K, recall_K, f1_K, ndcg_K, map_K = evaluation(tag_test, y_pred, TopK)
    print('acc: ', acc_K)
    print('precision: ', precision_K)
    print('recall: ', recall_K)
    print('f1: ', f1_K)
    print('ndcg: ', ndcg_K)
    print('map: ', map_K)
    np.savetxt(resultdir+method+'/y_pred'+str(i+1)+'.csv', y_pred, delimiter=",")
    save_result(acc_K, precision_K, recall_K, f1_K, ndcg_K, map_K)
wb.save(filename = resultdir+'evaluation2.xlsx')

#%% word embed/LSTM
#topK=10
#epoch=20
#softmax/mylossfunc
#acc:  0.6449332285938727
#precision:  0.12513747054202673
#recall:  0.49125425504058645
#f1:  0.18834801257355668

#softmax/BCE
#acc:  0.3974862529457973
#precision:  0.06857816182246662
#recall:  0.2483438072793925
#f1:  0.10207672713795352

#sigmoid/mylossfunc
#acc:  0.3534956794972506
#precision:  0.056402199528672425
#recall:  0.19914244566640482
#f1:  0.08375067420860129

#sigmoid/binarycrossentropy
#acc:  0.40062843676355064
#precision:  0.07038491751767478
#recall:  0.2526755321138668
#f1:  0.10466310691808589

#%% word embed/biLSTM
#topK=10
#epoch=20
#acc:  0.6449332285938727
#precision:  0.12513747054202673
#recall:  0.49125425504058645
#f1:  0.18834801257355668

#%% word embed/CNN1layer
#topK=10
#epoch=20
#acc:  0.6716417910447762
#precision:  0.13621366849960723
#recall:  0.5288847117794486
#f1:  0.20379186270757826

#%% 263個tag
#acc:  0.7151675485008818
#precision:  0.1416225749559083
#recall:  0.6035983735057809
#f1:  0.21682026986075506

#%% pretrain word2vec
#acc:  0.7892416225749559
#precision:  0.1564373897707231
#recall:  0.6785182385711486
#f1:  0.24027070768843503

#%% pretrain word2vec+selfattention / no pretrain model
#acc:  0.8774250440917107
#precision:  0.1734567901234568
#recall:  0.7840381008370427
#f1:  0.268484469955263