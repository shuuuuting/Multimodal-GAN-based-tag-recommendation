#%%
'''
from gensim.models import word2vec
embedding_size = 200
with open(datadir+'gensim_text.txt', 'a', encoding='utf-8') as file:
    for row in text_tr_pd.index:
        file.writelines(' '.join(text_tr_pd["content_jieba"][row]) + '\n' )

sentences = word2vec.LineSentence(datadir+'gensim_text.txt')
word2vec_model = word2vec.Word2Vec(sentences, size=embedding_size, min_count=1)
word2vec_model.save('word2vec_wiki.model')
'''

#%%
'''
embedding_matrix = np.zeros((len(model.wv.vocab.items()) + 1, model.vector_size))
word2idx = {}

vocab_list = [(word, model.wv[word]) for word, _ in model.wv.vocab.items()]
for i, vocab in enumerate(vocab_list):
    word, vec = vocab
    embedding_matrix[i + 1] = vec
    word2idx[word] = i + 1

def text_to_index(corpus):
    new_corpus = []
    for doc in corpus:
        new_doc = []
        for word in doc:
            try:
                new_doc.append(word2idx[word])
            except:
                new_doc.append(0)
        new_corpus.append(new_doc)
    return np.array(new_corpus)

pre_x_train = []
for i in text_tr_pd.index:  
        pre_x_train.append(text_tr_pd['content_jieba'][i])

import tensorflow as tf
import keras
from keras.preprocessing.sequence import pad_sequences

pre_x_train = text_to_index(pre_x_train)
pre_x_train = pad_sequences(pre_x_train, maxlen=50)
'''
#%%
'''
from numpy import genfromtxt
pre_x_train = genfromtxt(datadir+'pre_x_train.csv', delimiter=',')
pre_x_test = genfromtxt(datadir+'pre_x_test.csv', delimiter=',')
'''

#%%
from gensim.models import word2vec
w2vmodel = word2vec.Word2Vec.load('word2vec_wiki.model')

#%%
from keras.preprocessing.sequence import pad_sequences
embedding_train = []
i = 0
for docidx in text_tr_pd.index:
    embedding_train.append([])
    for word in text_tr_pd.content_jieba[docidx]:
        try:
            embedding_train[i].append(w2vmodel.wv[word])
        except:
            embedding_train[i].append(np.zeros(200))
    i+=1
embedding_train = pad_sequences(embedding_train,dtype='float32',maxlen=50) #(_,50,200)

embedding_test = []
i = 0
for docidx in text_te_pd.index:
    embedding_test.append([])
    for word in text_te_pd.content_jieba[docidx]:
        try:
            embedding_test[i].append(w2vmodel.wv[word])
        except:
            embedding_test[i].append(np.zeros(200))
    i+=1
embedding_test = pad_sequences(embedding_test,dtype='float32',maxlen=50)

#%%
from keras.models import Sequential
from keras.models import Model
from keras.layers.core import Activation, Flatten
from keras.layers.recurrent import LSTM
from keras.layers import Bidirectional
from keras.layers.convolutional import MaxPooling1D,Conv1D
from keras.layers import Input, Dense, Embedding, Dropout, GlobalMaxPooling1D, GlobalAveragePooling1D, Average, Add
from keras.optimizers import Adam
from keras_self_attention import SeqSelfAttention

embedding_size = 200 
seq_length = 50
TopK = 5
batch_size = 64
optimizer = Adam(lr=0.0005, decay=0.000001)
num_tags = tag_train.shape[1]
def textcnn():
    inp = Input(shape=(seq_length,embedding_size)) #+title_length
    #inp = w2vmodel.wv.get_keras_embedding(train_embeddings=True)(inpu)
    conv1 = Conv1D(embedding_size, 3, padding='same', activation='relu')(inp)
    #conv1 = SeqSelfAttention(attention_activation='sigmoid')(conv1)
    #conv1 = MaxPooling1D()(conv1)
    
    conv2 = Conv1D(embedding_size, 4, padding='same', activation='relu')(inp)
    #conv2 = SeqSelfAttention(attention_activation='sigmoid')(conv2)
    #conv2 = MaxPooling1D()(conv2)
    
    conv3 = Conv1D(embedding_size, 5, padding='same', activation='relu')(inp)
    #conv3 = SeqSelfAttention(attention_activation='sigmoid')(conv3)
    #conv3 = MaxPooling1D()(conv3)
    
    #cnn = concatenate([conv1, conv2, conv3], axis=-1, name='tcnn')
    cnn = Add(name='tcnn')([conv1, conv2, conv3])
    cnn = SeqSelfAttention(attention_activation='sigmoid')(cnn)
    cnn.set_shape((cnn.shape[0],seq_length,embedding_size)) #+title_length
    #cnn = AveragePooling1D(pool_size=seq_length//2, padding="same")(cnn) 
    flat = Flatten()(cnn)
    #flat = GlobalAveragePooling1D()(cnn)
    #x = Dense(1000, activation='relu')(flat)
    x = Dropout(0.2)(flat)
    #x = Dense(500, activation='relu')(x)
    #x = Dropout(0.5)(x)
    x = Dense(num_tags, activation='sigmoid')(x)
    model = Model(inputs=inp, outputs=x)
    model.compile(loss='BinaryCrossentropy', optimizer='adam', metrics=['accuracy'])
    model.summary()
    return model
model = textcnn()
'''
model = Sequential()
model.add(Conv1D(embedding_size, 3, padding='same', activation='relu', name='tcnn'))
model.add(SeqSelfAttention(attention_activation='sigmoid'))
model.add(MaxPooling1D())
#model.add(Flatten())
#model.add(GlobalMaxPooling1D())
model.add(GlobalAveragePooling1D())
model.add(Dense(600, activation='relu'))
model.add(Dense(num_tags, activation='softmax'))
model.compile(loss='binary_crossentropy', optimizer=optimizer, metrics=['accuracy'])
model.fit(embedding_train, tag_train.astype(np.float32), epochs=50, batch_size=batch_size)
model.save('pretrained_tCNN.h5')
'''
#%%
from openpyxl import load_workbook
from keras.callbacks import EarlyStopping
#text_title_train = np.concatenate((embedding_train,embedding_title_train),axis=1)
#text_title_test = np.concatenate((embedding_test,embedding_title_test),axis=1)
method = 'textCNN'
wb = load_workbook(resultdir+'evaluation.xlsx')
ws = wb.create_sheet(title=method)
ws.append(['acc', 'precision', 'recall', 'f1', 'ndcg', 'map'])
if not os.path.exists(resultdir+method):
    os.makedirs(resultdir+method)

early_stopping = EarlyStopping(monitor='val_loss', patience=10, verbose=1, mode='min', restore_best_weights=True) #持續5個epoch沒下降就停
round = 10
for i in range(round):
    print('round', i+1)
    model.fit(embedding_train, tag_train.astype(np.float32), epochs=50, batch_size=batch_size,)
            #validation_data=(embedding_test, tag_test), callbacks=[early_stopping])
    #model.save('pretrained_tCNN.h5')
    y_pred = model.predict(embedding_test)
    acc_K, precision_K, recall_K, f1_K, ndcg_K, map_K = evaluation(tag_test, y_pred, TopK)
    print('acc: ', acc_K)
    print('precision: ', precision_K)
    print('recall: ', recall_K)
    print('f1: ', f1_K)
    print('ndcg: ', ndcg_K)
    print('map: ', map_K)
    np.savetxt(resultdir+method+'/y_pred'+str(i+1)+'.csv', y_pred, delimiter=",")
    save_result(acc_K, precision_K, recall_K, f1_K, ndcg_K, map_K)
wb.save(filename = resultdir+'evaluation.xlsx')

##### word2vec pretrain by wiki #####
#%%
#acc:  0.7283950617283951
#precision:  0.13580246913580246
#recall:  0.5945091822737325
#f1:  0.2086098400155602

#%% flatten -> GlobalMaxPooling
#acc:  0.7768959435626103
#precision:  0.14991181657848324
#recall:  0.6555859998320316
#f1:  0.23043850036065053

#%% flatten -> GlobalAveragePooling
#acc:  0.843915343915344
#precision:  0.16331569664902998
#recall:  0.7336430251112791
#f1:  0.2524893453612309

#%% flatten -> GlobalMaxPooling
#   add self-attention after conv1d
#acc:  0.8465608465608465
#precision:  0.16631393298059965
#recall:  0.7430177626606198
#f1:  0.2568822101649268

#%% flatten -> GlobalAveragePooling
#   add self-attention after conv1d
#acc:  0.8518518518518519
#precision:  0.16666666666666666
#recall:  0.7473401500517902
#f1:  0.2575272058156147

#%%
from keras.preprocessing.text import Tokenizer
t = Tokenizer()
alldocs = list(text_tr_pd.sentence)
alldocs.extend(list(text_te_pd.sentence))
t.fit_on_texts(alldocs)
vocab_size = len(t.word_index) + 1
encoded_docs = t.texts_to_sequences(alldocs)
padded_docs = pad_sequences(encoded_docs,dtype='float32',maxlen=50)
embedding_matrix = np.zeros((vocab_size, w2vmodel.vector_size))
word2idx = {}
for word, i in t.word_index.items():
    if word in w2vmodel.wv.vocab:
        embedding_matrix[i] = w2vmodel.wv[word]
        word2idx[word] = i
pad_train = padded_docs[:4633,:]
pad_test = padded_docs[4633:,:]
embedding_layer = Embedding(embedding_matrix.shape[0],
                            embedding_matrix.shape[1],
                            weights=[embedding_matrix],
                            input_length=seq_length)
#acc:  0.7592592592592593
#precision:  0.14770723104056438
#recall:  0.6410787771898883
#f1:  0.2267760367853329
#%%
# get_keras_embedding
# input: pre_x_train / pre_x_test
#acc:  0.49382716049382713
#precision:  0.09497354497354497
#recall:  0.36991195655216813
#f1:  0.14292812217362155
