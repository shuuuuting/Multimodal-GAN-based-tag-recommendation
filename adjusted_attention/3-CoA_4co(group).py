#%%----------------------------------------------------------------------------------------
#seldDef.py
#------------------------------------------------------------------------------------------
import keras.backend as K
from keras.engine.topology import Layer, InputSpec
import keras
class coAttention_alt(Layer):
    """
    alternative co-attention
    inputs: [image feature tensor, hidden text feature tensor]
    output: co-Attention feature of image and text
    input dimensions:[(batchSize, num_region, CNN_dimension),
                    (batchSize, seq_length, CNN_dimension)]
    output dimension: batch_size*CNN_dimension
    """
    def __init__(self, dim_k, name='co_attn_layer'):
        super(coAttention_alt, self).__init__(name=name)
        self.dim_k = dim_k  # internal tensor dimension
        # self.input_spec = InputSpec(min_ndim=3)
        self.supports_masking = True

    def build(self, input_shape): #build:定義權重的function
        if not isinstance(input_shape, list): #input_shape要是list
            raise ValueError('A Co-Attention_alt layer should be called '
                             'on a list of inputs.')
        if len(input_shape) != 4: #input應該包含4個元件:img,text,group,title
            raise ValueError('A Co-Attention_alt layer should be called on a list of 4 inputs.'
                             'Got '+str(len(input_shape))+'inputs.')
        # print(input_shape)
        self.num_imgRegion = input_shape[0][1]
        self.seq_len = input_shape[1][1]
        self.group_len = input_shape[2][1]
        self.output_dim = input_shape[0][2] #200
        self.title_len = input_shape[3][1] #10

        self.w_Dense_Vit_0 = self.add_weight(name='w_Dense_Vit_0',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Vc_0 = self.add_weight(name='w_Dense_Vc_0',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Pc_0 = self.add_weight(name='w_Dense_Pc_0',
                                            shape=(2*self.dim_k, 1),
                                            initializer='random_normal',
                                            trainable=True)
        self.b_Dense_Pc_0 = self.add_weight(name='b_Dense_Pc_0',
                                            shape=(self.group_len,), 
                                            initializer='zeros',
                                            trainable=True)

        self.w_Dense_Ve_0 = self.add_weight(name='w_Dense_Ve_0',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Pe_0 = self.add_weight(name='w_Dense_Pe_0',
                                            shape=(2*self.dim_k, 1),
                                            initializer='random_normal',
                                            trainable=True)
        self.b_Dense_Pe_0 = self.add_weight(name='b_Dense_Pe_0',
                                            shape=(self.title_len,), 
                                            initializer='zeros',
                                            trainable=True)

        self.w_Dense_Vi_1 = self.add_weight(name='w_Dense_Vi_1',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Vtce_1 = self.add_weight(name='w_Dense_Vtce_1',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Pi_1 = self.add_weight(name='w_Dense_Pi_1',
                                            shape=(2*self.dim_k, 1),
                                            initializer='random_normal',
                                            trainable=True)
        self.b_Dense_Pi_1 = self.add_weight(name='b_Dense_Pi_1',
                                            shape=(self.num_imgRegion,),
                                            initializer='zeros',
                                            trainable=True)

        self.w_Dense_Vice_2 = self.add_weight(name='w_Dense_Vice_2',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Vt_2 = self.add_weight(name='w_Dense_Vt_2',
                                            shape=(self.output_dim, self.dim_k),
                                            initializer='random_normal',
                                            trainable=True)
        self.w_Dense_Pt_2 = self.add_weight(name='w_Dense_Pt_2',
                                            shape=(2*self.dim_k, 1),
                                            initializer='random_normal',
                                            trainable=True)
        self.b_Dense_Pt_2 = self.add_weight(name='b_Dense_Pt_2',
                                            shape=(self.seq_len,),
                                            initializer='zeros',
                                            trainable=True) 

        super(coAttention_alt, self).build(input_shape)  # Be sure to call this somewhere! 一定要在最後調用它!

    def call(self, x, mask=None): #call:編寫layer邏輯的function，執行forward propagation
        ifeature_h = x[0]
        tfeature_h = x[1]
        cfeature = x[2]
        efeature = x[3]
        output_dim = self.output_dim
        num_imgRegion = self.num_imgRegion
        dim_k = self.dim_k
        seq_len = self.seq_len
        ifeature = K.mean(ifeature_h, axis=1) #(batchSize, output_dim=200)
        tfeature = K.mean(tfeature_h, axis=1) #(batchSize, output_dim=200)

        alpha = 0.5
        beta = 0.5

        # phase 0: i+t guide popular and multi-gCNN
        w_Vit_0 = K.repeat(K.dot(ifeature+tfeature, self.w_Dense_Vit_0), self.group_len)    # shape=(batchSize,seq_len,dim_k)
        w_Vc_0 = K.dot(K.reshape(cfeature, [-1, output_dim]), self.w_Dense_Vc_0)   # shape=((batchSize*seq_len),dim_k)
        w_Vc_0 = K.reshape(w_Vc_0, (-1, self.group_len, dim_k))    # shape= (batchSize,seq_len,dim_k)
        Vit_Vc_0 = K.concatenate([w_Vit_0, w_Vc_0], axis=-1)    # shape=(batchSize,seq_len,2*dim_k)
        Hc = K.tanh(Vit_Vc_0)
        Hc_b = K.squeeze(K.dot(Hc, self.w_Dense_Pc_0), axis=-1) + self.b_Dense_Pc_0   # shape=(batch_size,seq_len)
        Pc = K.softmax(Hc_b)
        Pc = K.permute_dimensions(K.repeat(Pc, output_dim), (0, 2, 1))    # shape=(batchSize,seq_len,output_dim)
        Pc_Vit = Pc*cfeature    # shape=(batchSize,seq_len,output_dim)*(batchSize,seq_len,output_dim)
        Vc = K.sum(Pc_Vit, axis=1)   # shape=(batchSize,output_dim)

        w_Vit_0 = K.repeat(K.dot(ifeature+tfeature, self.w_Dense_Vit_0), self.title_len)    # shape=(batchSize,title_len,dim_k)
        w_Ve_0 = K.dot(K.reshape(efeature, [-1, output_dim]), self.w_Dense_Ve_0)   # shape=((batchSize*title_len),dim_k)
        w_Ve_0 = K.reshape(w_Ve_0, (-1, self.title_len, dim_k))    # shape= (batchSize,title_len,dim_k)
        Vit_Ve_0 = K.concatenate([w_Vit_0, w_Ve_0], axis=-1)    # shape=(batchSize,title_len,2*dim_k)
        He = K.tanh(Vit_Ve_0)
        He_b = K.squeeze(K.dot(He, self.w_Dense_Pe_0), axis=-1) + self.b_Dense_Pe_0   # shape=(batch_size,title_len)
        Pe = K.softmax(He_b)
        Pe = K.permute_dimensions(K.repeat(Pe, output_dim), (0, 2, 1))    # shape=(batchSize,title_len,output_dim)
        Pe_Vit = Pe*efeature    # shape=(batchSize,title_len,output_dim)*(batchSize,title_len,output_dim)
        Ve = K.sum(Pe_Vit, axis=1)   # shape=(batchSize,output_dim)

        # phase 1: text,Ve,Vc guide image 
        w_Vi_1 = K.dot(K.reshape(ifeature_h, [-1, output_dim]), self.w_Dense_Vi_1) # shape=((batchSize*num_imgRegion),dim_k)
        w_Vi_1 = K.reshape(w_Vi_1, [-1, num_imgRegion, dim_k])  # shape=(batchSize,num_imgRegion,dim_k)
        w_Vtce_1 = K.repeat(K.dot(tfeature+alpha*Vc+beta*Ve, self.w_Dense_Vtce_1), num_imgRegion)  # shape=(batchSize,num_imgRegion,dim_k) #未repeat前的dim是(batchSize,dim_k)
        Vi_Vtce_1 = K.concatenate([w_Vi_1, w_Vtce_1], axis=-1)  # shape=(batchSize,num_imgRegion,2*dim_k)
        Hi = K.tanh(Vi_Vtce_1)
        Hi_w_b = K.squeeze(K.dot(Hi, self.w_Dense_Pi_1), axis=-1) + self.b_Dense_Pi_1  # shape=(batchSize,num_imgRegion) #axis是要丟棄的軸 #squeeze是用來刪掉維數是1的維度
        Pi = K.softmax(Hi_w_b) # shape=(batchSize,num_imgRegion)
        Pi = K.permute_dimensions(K.repeat(Pi, output_dim), (0, 2, 1))  # shape=(batchSize,num_imgRegion,output_dim)
        Pi_Vi = Pi*ifeature_h # shape=(batchSize,num_imgRegion,output_dim)*(batchSize,num_imgRegion,output_dim)
        Vi = K.sum(Pi_Vi, axis=1)  # shape=(batchSize,output_dim)
        
        # phase 2: Vi,Ve,Vc guide text 
        w_Vice_2 = K.repeat(K.dot(Vi+alpha*Vc+beta*Ve, self.w_Dense_Vice_2), seq_len)    # shape=(batchSize,seq_len,dim_k)
        w_Vt_2 = K.dot(K.reshape(tfeature_h, [-1, output_dim]), self.w_Dense_Vt_2)   # shape=((batchSize*seq_len),dim_k)
        w_Vt_2 = K.reshape(w_Vt_2, (-1, seq_len, dim_k))    # shape= (batchSize,seq_len,dim_k)
        Vice_Vt_2 = K.concatenate([w_Vice_2, w_Vt_2], axis=-1)    # shape=(batchSize,seq_len,2*dim_k)
        Ht = K.tanh(Vice_Vt_2)
        Ht_b = K.squeeze(K.dot(Ht, self.w_Dense_Pt_2), axis=-1) + self.b_Dense_Pt_2   # shape=(batch_size,seq_len)
        Pt = K.softmax(Ht_b)
        Pt = K.permute_dimensions(K.repeat(Pt, output_dim), (0, 2, 1))    # shape=(batchSize,seq_len,output_dim)
        Pt_Vt = Pt*tfeature_h # shape=(batchSize,seq_len,output_dim)*(batchSize,seq_len,output_dim)
        Vt = K.sum(Pt_Vt, axis=1)   # shape=(batchSize,output_dim)
        
        return Vi + Vt + alpha*Vc + beta*Ve

    def compute_output_shape(self, input_shape):
        output_shape = (input_shape[0][0], input_shape[0][-1]) #(batch_size,CNN_dimension)
        return output_shape

    def get_config(self):
        return super(coAttention_alt, self).get_config()

def myLossFunc(y_true, y_pred):
    probs_log = -K.log(y_pred)
    loss = K.mean(K.sum(probs_log*y_true, axis=-1))
    # loss = K.mean(K.sum(K.clip(probs_log * y_true, -1e40, 100), axis=-1))
    return loss

#%%----------------------------------------------------------------------------------------
#co-attention.py 
#------------------------------------------------------------------------------------------
from keras.models import Model
from keras.layers.core import Activation, Flatten, Reshape, RepeatVector
from keras.layers.recurrent import LSTM
from keras.layers import Bidirectional
from keras.layers.convolutional import AveragePooling1D,Conv1D
from keras.layers.wrappers import TimeDistributed
from keras.layers import Input, Dense, Embedding, Dropout, Lambda

num_tags = tag_train.shape[1]
num_words = vocab_size
#index_from = 3
seq_length = max_length
title_length = 10
batch_size = 64
embedding_size = 200
hidden_size = 100
attention_size = 200
dim_k = 100
num_region = 7*7
drop_rate = 0.5
TopK = 5

def imageFeature(inputs):
    features = Reshape(target_shape=(num_region, 512))(inputs)
    features = Dense(embedding_size, activation="tanh", use_bias=False)(features) #single layer to convert each img vector into a new same dim vector as text feature vector
    features = SeqSelfAttention(attention_activation='sigmoid')(features)
    features_pooling = AveragePooling1D(pool_size=num_region, padding="same")(features)
    features_pooling = Lambda(lambda x: K.squeeze(x, axis=1))(features_pooling)

    return features, features_pooling

def textFeature(X):
    #embeddings = Embedding(input_dim=num_words, output_dim=embedding_size,
    #                       mask_zero=True, input_length=seq_length)(X) #seq_length:代表取前幾個熱門字
    tFeature1 = Conv1D(embedding_size, 3, padding='same', strides=1, activation='relu')(X)
    tFeature2 = Conv1D(embedding_size, 4, padding='same', strides=1, activation='relu')(X)
    tFeature3 = Conv1D(embedding_size, 5, padding='same', strides=1, activation='relu')(X)
    tFeature = Add(name='tcnn')([tFeature1, tFeature2, tFeature3])
    tFeature = SeqSelfAttention(attention_activation='sigmoid')(tFeature)

    return tFeature

def groupFeature(X):
    #X = Embedding(input_dim=num_words, output_dim=embedding_size,
    #                       mask_zero=True, input_length=seq_length)(X) #seq_length:一次輸入带有的詞彙個數
    #gFeature = LSTM(units=embedding_size, return_sequences=True)(embeddings)
    #gFeature = Bidirectional(LSTM(units=embedding_size, return_sequences=True), merge_mode='sum')(embeddings)
    gFeature1 = Conv1D(embedding_size, 3, padding='same', strides=1, activation='relu')(X)
    gFeature2 = Conv1D(embedding_size, 4, padding='same', strides=1, activation='relu')(X)
    gFeature3 = Conv1D(embedding_size, 5, padding='same', strides=1, activation='relu')(X)
    gFeature = Add()([gFeature1, gFeature2, gFeature3])
    gFeature = SeqSelfAttention(attention_activation='sigmoid')(gFeature)
    return gFeature

def titleFeature(X):
    eFeature1 = Conv1D(embedding_size, 3, padding='same', strides=1, activation='relu')(X)
    eFeature2 = Conv1D(embedding_size, 4, padding='same', strides=1, activation='relu')(X)
    eFeature3 = Conv1D(embedding_size, 5, padding='same', strides=1, activation='relu')(X)
    eFeature = Add()([eFeature1, eFeature2, eFeature3])
    eFeature = SeqSelfAttention(attention_activation='sigmoid')(eFeature)
    return eFeature

def modelDef():
    inputs_img = Input(shape=(7, 7, 512))
    #inputs_text = Input(shape=(seq_length,embedding_size))
    inputs_text = Input(shape=(seq_length,embedding_size)) #concate w/ title
    inputs_channel = Input(shape=(seq_length,embedding_size)) #embedding
    inputs_title = Input(shape=(seq_length,embedding_size)) #embedding

    iFeature, iFeature_pooling = imageFeature(inputs_img)
    tFeature = textFeature(inputs_text)
    gFeature = groupFeature(inputs_channel) #embedding
    eFeature = titleFeature(inputs_title) #embedding

    iFeature.set_shape((inputs_img.shape[0],num_region,embedding_size))
    #tFeature.set_shape((inputs_text.shape[0],seq_length,embedding_size))
    tFeature.set_shape((inputs_text.shape[0],seq_length,embedding_size)) #concate w/ title
    gFeature.set_shape((inputs_channel.shape[0],seq_length,embedding_size))
    eFeature.set_shape((inputs_title.shape[0],seq_length,embedding_size))

    co_feature = coAttention_alt(dim_k=dim_k)([iFeature, tFeature, gFeature, eFeature])
    dropout = Dropout(drop_rate)(co_feature)
    #all_feature = attn(dropout, gFeature)
    Softmax = Dense(num_tags, activation="softmax", use_bias=True)(dropout)
    Sigmoid = Dense(num_tags, activation="sigmoid", use_bias=True)(dropout)
    model = Model(inputs=[inputs_img, inputs_text, inputs_channel, inputs_title],
                  outputs=[Sigmoid])
    # adam = optimizers.adam(lr=0.001, beta_1=0.9, beta_2=0.999, epsilon=1e-08, decay=0.00001)
    model.compile(optimizer="adam", loss='BinaryCrossentropy')
    #model.compile(optimizer="adam", loss=myLossFunc)
    return model

#%%----------------------------------------------------------------------------------------
#run co-attention.py 
#------------------------------------------------------------------------------------------
from openpyxl import load_workbook
from keras.callbacks import EarlyStopping
method = 'CoA_group'
wb = load_workbook(resultdir+'evaluation2.xlsx')
ws = wb.create_sheet(title=method)
ws.append(['acc', 'precision', 'recall', 'f1', 'ndcg', 'map'])
if not os.path.exists(resultdir+method):
    os.makedirs(resultdir+method)

early_stopping = EarlyStopping(monitor='val_loss', patience=10, verbose=1, mode='min', restore_best_weights=True) #持續5個epoch沒下降就停
round = 10
for i in range(round):
    print('round', i+1)
    model = modelDef()
    history = model.fit(x=[image_train, embedding_train, lda_train, lda_tr_arr], 
                        y=tag_train.astype(np.float32),
                        batch_size=batch_size,
                        epochs=50,
                        verbose=1,)
                        #validation_data=([image_test, embedding_test, channel_test, embedding_title_test], tag_test.astype(np.float32)), 
                        #callbacks=[early_stopping])
    y_pred = model.predict(x=[image_test, embedding_test, lda_test, lda_te_arr]) 
    acc_K, precision_K, recall_K, f1_K, ndcg_K, map_K = evaluation(tag_test, y_pred, TopK)
    print('acc: ', acc_K)
    print('precision: ', precision_K)
    print('recall: ', recall_K)
    print('f1: ', f1_K)
    print('ndcg: ', ndcg_K)
    print('map: ', map_K)
    np.savetxt(resultdir+method+'/y_pred'+str(i+1)+'.csv', y_pred, delimiter=",")
    save_result(acc_K, precision_K, recall_K, f1_K, ndcg_K, map_K)
wb.save(filename = resultdir+'evaluation2.xlsx')

#%% phase 2 i+t guide channel
# + phase 3 i+t guide title 
#alpha=0.5/beta=0.5
#acc:  0.8209876543209876
#precision:  0.16349206349206347
#recall:  0.7251690182245738
#f1:  0.2521570785801624
